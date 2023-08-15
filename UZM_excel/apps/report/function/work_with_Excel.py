import os
from math import sqrt
from typing import NoReturn

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill

from .graffic import get_graphics

file_dir = os.getcwd() + "\\Files"


def cell_copy(sheet, *, row: int, col: str) -> list:
    """Копирование ячеек из экселя в list"""
    data_List = []
    # print(f'Считывание до строки - {sheet.max_row}')
    for column in sheet[col]:
        if column.row in range(row, sheet.max_row + 1):
            if column.value == "N/A":
                data_List.append(0)
            else:
                try:
                    # print(float(column.value))
                    data_List.append(float(column.value))
                except Exception as e:
                    break
    return data_List


def excel_open(column: dict) -> dict:
    """
        Чтение excel файлов на основе переданных параметров - column - это словарь с колонками и данными о файле
            На входе:
                'Глубина'  - имя столбца с глубиной
                'Угол'     - имя столбца с угом
                'Азимут'   - имя столбца с азимутом (опционально)
                'Имя файла'- имя файла для считывания из папки UZM_excel\Files\Report_input
                'Строка'   - строка с которой считываем данные

            На выходе: (если найдем имя столбца в файле)
                'Глубина'  - Данные [1,2,3..,4]
                'Угол'     - Данные [1,2,3..,4]
                'Азимут'   - Данные [1,2,3..,4]
    """
    result = {}
    filename = column['Имя файла']

    try:
        file_folder = file_dir + '\\Report_input\\' + filename
        excel_file = openpyxl.load_workbook(file_folder, data_only=True)
        excel_sheet = excel_file[column['Лист']]
    except Exception:
        print(Exception)
        print(f'Ошибка при работе с файлом: {filename}\n'
              f'в папке: {file_folder}\n'
              f'Данные на входе: {column}\n'
              f'Открывал лист с именем: {column["Лист"]}')

    for key, value in column.items():
        if key == "Лист" or key == 'Строка':
            break  # обрабатываем данные только с индексами колонок
        else:
            result[key] = cell_copy(excel_sheet, row=int(column['Строка']), col=value)

    excel_file.close()
    return result


"""
    Функции для записи данных в excel
"""


def write_data_in_Excel(all_data: dict, filename: str, short_type: int, Run: object) -> str:
    """
    Записываем замеры из all_data в шаблон
    filename - имя шаблона в папке Шаблон
    short_type - сокращенный файл (0) или полный (1)
    Well- скважина

    На выходе получаем имя файла, которое отдаем на выдачу и параметры отходов
    """
    well = Run.section.wellbore.well_name
    wellbore = Run.section.wellbore
    # построение графика
    other_data, waste_word = get_graphics(all_data, wellbore)  # other_data - тут лежат TVD, угол, азимут

    Field = well.pad_name.field
    report_type = {'шаблон': 'единый'}
    if Field.field_name == 'Северо-Комсомольское':  # используем шаблон с картографическим азимутом
        filename = 'Северо-Комсомольское_отчёт.xlsx'
        report_type['шаблон'] = 'северокомсомольский'
    elif Field.field_name == 'Самотлорское':  # шаблон с номером рейса, без абсолютной отметки
        filename = 'Самотлорское_отчёт.xlsx'
        report_type['шаблон'] = 'самотлорское'

    file_folder = file_dir + '\\Шаблон\\' + filename
    excel_file = openpyxl.load_workbook(file_folder)
    # print(all_data['Статические замеры ННБ'])  для отладки
    # print(all_data['Статические замеры ИГИРГИ'])
    hor, ver, common = write_data(excel_file,  # горизонтальные, вертикальные, общие отходы
                                  all_data['Статические замеры ННБ'],
                                  all_data['Статические замеры ИГИРГИ'],
                                  other_data,
                                  Run,
                                  report_type)

    waste = {'hor': hor,
             'ver': ver,
             'common': common,
             'word': waste_word,
             }
    if short_type == 1:  # полный формат с доп данными
        nnb_dynamic(excel_file, all_data['Динамические замеры ННБ'])
        write_dynamic_igirgi(excel_file,
                             all_data['Динамические замеры ИГИРГИ'],
                             all_data['Статические замеры ИГИРГИ'])

    # # # Отладочная информация в отчёте
    # excel_file.create_sheet('Отладка', 3)
    # excel_sheet = excel_file['Отладка']
    # j = 1
    # for key, column_dict in all_data.items():
    #     excel_sheet.cell(row=1, column=j).value = key
    #     for column_name, data_list in column_dict.items():
    #         excel_sheet.cell(row=2, column=j).value = column_name
    #         i = 3
    #         for value in data_list:
    #             excel_sheet.cell(row=i, column=j).value = value
    #             i += 1
    #         j += 1

    # добавляем график
    grafic = Image(file_dir + f'\\Report_out\\{wellbore}.png')
    excel_file['Проекции'].add_image(grafic, 'A1')

    report_name = get_excel_name(well,
                                 last_depth=all_data['Статические замеры ИГИРГИ']['Глубина'][-1],
                                 departure_ver=ver,
                                 departure_horiz=hor, )
    excel_file.save(file_dir + '\\Report_out\\' + report_name)
    return report_name, waste


def write_dynamic_igirgi(excel_file: openpyxl.workbook.workbook.Workbook,
                         dynamic_data: dict,
                         static_data: dict):
    """
    Записывает динамические замеры в эксель, помечаем желтым статические данные
    """
    # excel_file.create_sheet('Динамика ИГиРГИ', 2)
    copy_hat(excel_file, 'Динамика ИГиРГИ')

    excel_sheet = excel_file['Динамика ИГиРГИ']
    row = 17
    for meas in zip(dynamic_data['Глубина'], dynamic_data['Угол'], dynamic_data['Азимут']):
        excel_sheet.cell(row=row, column=1).value = meas[0]
        excel_sheet.cell(row=row, column=2).value = meas[1]
        excel_sheet.cell(row=row, column=3).value = meas[2]
        if meas[0] in static_data['Глубина']:
            excel_sheet.cell(row=row, column=1).fill = PatternFill('solid', fgColor='EDFF21')
            excel_sheet.cell(row=row, column=2).fill = PatternFill('solid', fgColor='EDFF21')
            excel_sheet.cell(row=row, column=3).fill = PatternFill('solid', fgColor='EDFF21')
        row += 1


def nnb_dynamic(excel: openpyxl.workbook.workbook.Workbook, data: dict) -> NoReturn:
    """Записывает в отчет динамику ННБ"""
    # print(excel)
    # excel.create_sheet('Динамика ННБ', 2)
    copy_hat(excel, 'Динамика ННБ')
    excel_sheet = excel['Динамика ННБ']
    for key, array in data.items():
        row = 17
        if key == 'Глубина':
            col = 1
        if key == 'Угол':
            col = 2
        if key == 'Азимут':
            col = 3
        for element in array:
            excel_sheet.cell(row=row, column=col).value = element
            row += 1


def write_data(excel: openpyxl.workbook.workbook.Workbook,
               nnb: dict, igirgi: dict, other: dict,
               Run: object, report_type: dict) -> float:
    """ Первая страница запись (Данные)
        Возвращаем последние посчитанные отходы на для названия файла
    """
    Well = Run.section.wellbore.well_name
    excel_sheet = excel['Данные']

    write_hat(excel['Данные'], Well)  # шапка для страницы
    if report_type['шаблон'] != 'единый':  # обработчик нестандартных шаблонов
        if report_type['шаблон'] == 'северокомсомольский':
            sevcom_hat(excel['Данные'], Well)
            return sevcom_data(excel_sheet, nnb, igirgi, other, Well)
        elif report_type['шаблон'] == 'самотлорское':
            samotlor_hat(excel['Данные'], Well)
            return samotlor_data(excel_sheet, nnb, igirgi, other, Well)

    # стандартное заполнение отчёта (можно вынести в отдельную функцию)
    table_hat(excel['Данные'], Well)  # шапка для таблицы
    row = 18
    count = 0  # первую стрчоку не выводим
    # запись в ячейки
    for meas in zip(igirgi['Глубина'], igirgi['Угол'], igirgi['Азимут'], nnb['Угол'], nnb['Азимут'],
                    other['igirgi_TVDSS'], other['igirgi_delta_x'], other['igirgi_delta_y'], other['igirgi_TVD'],
                    other['nnb_delta_x'], other['nnb_delta_y'], other['nnb_TVD'], nnb['Комментарий']):
        if count == 0:
            count += 1
            continue
        # excel_sheet.cell(row=row, column=1).value = Run.run_number  # Номер рейса
        excel_sheet.cell(row=row, column=1).value = round(meas[0], 2)  # Глубина
        excel_sheet.cell(row=row, column=2).value = round(meas[1], 2)  # Зенитный угол
        excel_sheet.cell(row=row, column=3).value = round(meas[2], 2)  # Азимут картографический
        if Well.dec is None:
            excel_sheet.cell(row=row, column=4).value = '-'  # Азимут магнитный
        else:
            azimut = round(meas[2] - Well.total_correction, 2)
            excel_sheet.cell(row=row, column=4).value = azimut if azimut > 0 else azimut + 360  # Азимут магнитный

        excel_sheet.cell(row=row, column=5).value = round(meas[5], 2)  # Абсолютная отметка TVDSS
        excel_sheet.cell(row=row, column=6).value = round(meas[3], 2)  # ННБ зенитный угол
        excel_sheet.cell(row=row, column=7).value = round(meas[4], 2)  # ННБ азимут
        excel_sheet.cell(row=row, column=8).value = round(meas[1] - meas[3], 2)  # разница зенитный угол
        dif_Az = meas[2] - meas[4]  # разница азимут
        excel_sheet.cell(row=row, column=9).value = round((dif_Az if dif_Az <= 300 else dif_Az - 360), 2)
        # пошли отходы
        Ex = (Well.EX if Well.EX is not None else 0)
        Ny = (Well.NY if Well.NY is not None else 0)
        X_nnb = Ex + meas[10]
        Y_nnb = Ny + meas[9]

        X_igirgi = Ex + meas[7]
        Y_igigri = Ny + meas[6]

        excel_sheet.cell(row=row, column=10).value = round(  # отход по горизонтали
            sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2), 2)
        excel_sheet.cell(row=row, column=11).value = round(meas[11] - meas[8], 2)  # отход по вертикали
        excel_sheet.cell(row=row, column=12).value = round(  # отход общий
            sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2 +
                 (meas[11] - meas[8]) ** 2), 2)

        excel_sheet.cell(row=row, column=13).value = meas[12]

        if row == (17 + len(igirgi['Глубина']) - 1):  # забираем последнии значения отходов
            return round(sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2), 2), round(meas[11] - meas[8], 2), \
                round(sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2 + (meas[11] - meas[8]) ** 2), 2)
        row += 1


def sevcom_data(excel_sheet: openpyxl.workbook.workbook.Workbook,
                nnb: dict, igirgi: dict, other: dict,
                Well: object) -> float:
    """Нестандартный формат отчёта с доп колонкой на картографический азимут"""
    row = 18
    count = 0  # первую строку не выводим
    # запись в ячейки
    for meas in zip(igirgi['Глубина'], igirgi['Угол'], igirgi['Азимут'], nnb['Угол'], nnb['Азимут'],
                    other['igirgi_TVDSS'], other['igirgi_delta_x'], other['igirgi_delta_y'], other['igirgi_TVD'],
                    other['nnb_delta_x'], other['nnb_delta_y'], other['nnb_TVD'], nnb['Комментарий']):
        if count == 0:
            count += 1
            continue
        # excel_sheet.cell(row=row, column=1).value = Run.run_number  # Номер рейса
        excel_sheet.cell(row=row, column=1).value = round(meas[0], 2)  # Глубина
        excel_sheet.cell(row=row, column=2).value = round(meas[1], 2)  # Зенитный угол
        excel_sheet.cell(row=row, column=3).value = round(meas[2], 2)  # Азимут географический
        if Well.grid_convergence is None:  # Азимут картографический NEW
            excel_sheet.cell(row=row, column=4).value = '-'
        else:
            azimut = round(meas[2] - Well.grid_convergence, 2)
            excel_sheet.cell(row=row, column=4).value = azimut if azimut > 0 else azimut + 360

        if Well.dec is None:  # Азимут магнитный
            excel_sheet.cell(row=row, column=5).value = '-'
        else:
            azimut = round(meas[2] - Well.total_correction, 2)
            excel_sheet.cell(row=row, column=5).value = azimut if azimut > 0 else azimut + 360

        excel_sheet.cell(row=row, column=6).value = round(meas[5], 2)  # Абсолютная отметка TVDSS
        excel_sheet.cell(row=row, column=7).value = round(meas[3], 2)  # ННБ зенитный угол
        excel_sheet.cell(row=row, column=8).value = round(meas[4], 2)  # ННБ азимут
        excel_sheet.cell(row=row, column=9).value = round(meas[1] - meas[3], 2)  # разница зенитный угол
        dif_Az = meas[2] - meas[4]  # разница азимут
        excel_sheet.cell(row=row, column=9).value = round((dif_Az if dif_Az <= 300 else dif_Az - 360), 2)
        # пошли отходы
        Ex = (Well.EX if Well.EX is not None else 0)
        Ny = (Well.NY if Well.NY is not None else 0)
        X_nnb = Ex + meas[10]
        Y_nnb = Ny + meas[9]

        X_igirgi = Ex + meas[7]
        Y_igigri = Ny + meas[6]

        excel_sheet.cell(row=row, column=11).value = round(  # отход по горизонтали
            sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2), 2)
        excel_sheet.cell(row=row, column=12).value = round(meas[11] - meas[8], 2)  # отход по вертикали
        excel_sheet.cell(row=row, column=13).value = round(  # отход общий
            sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2 +
                 (meas[11] - meas[8]) ** 2), 2)

        excel_sheet.cell(row=row, column=14).value = meas[12]

        if row == (17 + len(igirgi['Глубина']) - 1):  # забираем последнии значения отходов
            return round(sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2), 2), round(meas[11] - meas[8], 2), \
                round(sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2 + (meas[11] - meas[8]) ** 2), 2)
        row += 1


def samotlor_data(excel_sheet: openpyxl.workbook.workbook.Workbook,
                  nnb: dict, igirgi: dict, other: dict,
                  Well: object) -> float:
    """Нестандартный формат отчёта с номером рейса и без абсолютной отметки"""
    row = 18
    count = 0  # первую строку не выводим
    # запись в ячейки
    for meas in zip(igirgi['Глубина'], igirgi['Угол'], igirgi['Азимут'], nnb['Угол'], nnb['Азимут'],
                    other['igirgi_TVDSS'], other['igirgi_delta_x'], other['igirgi_delta_y'], other['igirgi_TVD'],
                    other['nnb_delta_x'], other['nnb_delta_y'], other['nnb_TVD'], nnb['Комментарий'], nnb['Рейс']):
        if count == 0:
            count += 1
            continue
        excel_sheet.cell(row=row, column=1).value = (meas[13] if meas[13] != -1 else 'материнский ствол')  # Номер рейса
        excel_sheet.cell(row=row, column=2).value = round(meas[0], 2)  # Глубина
        excel_sheet.cell(row=row, column=3).value = round(meas[1], 2)  # Зенитный угол
        excel_sheet.cell(row=row, column=4).value = round(meas[2], 2)  # Азимут

        if Well.dec is None:  # Азимут магнитный
            excel_sheet.cell(row=row, column=5).value = '-'
        else:
            azimut = round(meas[2] - Well.total_correction, 2)
            excel_sheet.cell(row=row, column=5).value = azimut if azimut > 0 else azimut + 360

        excel_sheet.cell(row=row, column=6).value = round(meas[3], 2)  # ННБ зенитный угол
        excel_sheet.cell(row=row, column=7).value = round(meas[4], 2)  # ННБ азимут
        excel_sheet.cell(row=row, column=8).value = round(meas[1] - meas[3], 2)  # разница зенитный угол
        dif_Az = meas[2] - meas[4]  # разница азимут
        excel_sheet.cell(row=row, column=9).value = round((dif_Az if dif_Az <= 300 else dif_Az - 360), 2)

        # пошли отходы
        Ex = (Well.EX if Well.EX is not None else 0)
        Ny = (Well.NY if Well.NY is not None else 0)
        X_nnb = Ex + meas[10]
        Y_nnb = Ny + meas[9]

        X_igirgi = Ex + meas[7]
        Y_igigri = Ny + meas[6]

        excel_sheet.cell(row=row, column=10).value = round(  # отход по горизонтали
            sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2), 2)
        excel_sheet.cell(row=row, column=11).value = round(meas[11] - meas[8], 2)  # отход по вертикали
        excel_sheet.cell(row=row, column=12).value = round(  # отход общий
            sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2 +
                 (meas[11] - meas[8]) ** 2), 2)

        excel_sheet.cell(row=row, column=13).value = meas[12]

        if row == (17 + len(igirgi['Глубина']) - 1):  # забираем последнии значения отходов
            return round(sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2), 2), round(meas[11] - meas[8], 2), \
                round(sqrt((X_nnb - X_igirgi) ** 2 + (Y_nnb - Y_igigri) ** 2 + (meas[11] - meas[8]) ** 2), 2)
        row += 1


def sevcom_hat(excel_sheet: openpyxl.workbook.workbook.Workbook.worksheets, Well: object) -> NoReturn:
    """Шапка колонок таблицы для севкомоского отчёта"""
    excel_sheet['C16'].value = '-' if Well.north_direction is None else Well.get_north_direction()
    excel_sheet['H16'].value = '-' if Well.north_direction is None else Well.get_north_direction()
    excel_sheet['J16'].value = '-' if Well.north_direction is None else Well.get_north_direction()


def samotlor_hat(excel_sheet: openpyxl.workbook.workbook.Workbook.worksheets, Well: object) -> NoReturn:
    """Шапка колонок таблицы для самотлорского отчёта"""
    excel_sheet['D16'].value = '-' if Well.north_direction is None else Well.get_north_direction()
    excel_sheet['G16'].value = '-' if Well.north_direction is None else Well.get_north_direction()
    excel_sheet['I16'].value = '-' if Well.north_direction is None else Well.get_north_direction()


def table_hat(excel_sheet: openpyxl.workbook.workbook.Workbook.worksheets, Well: object) -> NoReturn:
    """Стандартная шапка колонок таблицы"""
    excel_sheet['C16'].value = '-' if Well.north_direction is None else Well.get_north_direction()
    excel_sheet['G16'].value = '-' if Well.north_direction is None else Well.get_north_direction()
    excel_sheet['I16'].value = '-' if Well.north_direction is None else Well.get_north_direction()


def write_hat(excel_sheet: openpyxl.workbook.workbook.Workbook.worksheets, Well: object) -> NoReturn:
    """
    Записываем шапку Excel страницы по данным модели Well
    """
    excel_sheet['C3'].value = Well.pad_name.field.field_name
    excel_sheet['C4'].value = Well.pad_name.pad_name
    excel_sheet['C5'].value = Well.well_name
    excel_sheet['C6'].value = '-' if Well.RKB is None else Well.RKB
    excel_sheet['C7'].value = '-' if Well.coordinate_system is None else Well.coordinate_system
    excel_sheet['C8'].value = '-' if Well.latitude is None else Well.latitude
    excel_sheet['C9'].value = '-' if Well.longtitude is None else Well.longtitude
    excel_sheet['C10'].value = '-' if Well.NY is None else Well.NY
    excel_sheet['C11'].value = '-' if Well.EX is None else Well.EX
    excel_sheet['H3'].value = '-' if Well.geomagnetic_model is None else Well.geomagnetic_model
    excel_sheet['H4'].value = '-' if Well.geomagnetic_date is None else Well.geomagnetic_date
    excel_sheet['H5'].value = '-' if Well.north_direction is None else Well.get_north_direction()
    excel_sheet['H6'].value = '-' if Well.btotal is None else Well.btotal
    excel_sheet['H7'].value = '-' if Well.dip is None else Well.dip
    excel_sheet['H8'].value = '-' if Well.dec is None else Well.dec
    excel_sheet['H9'].value = '-' if Well.grid_convergence is None else Well.grid_convergence
    excel_sheet['H10'].value = '-' if Well.total_correction is None else Well.total_correction
    excel_sheet['H11'].value = '-' if Well.gtotal is None else Well.gtotal


def copy_hat(excel: openpyxl.workbook.workbook.Workbook, sheet_name: str) -> NoReturn:
    """
    Чтобы повторно не обращаться к БД просто копируем шапку с страницы Данные
    """
    from_sheet = excel['Данные']
    to_sheet = excel[sheet_name]
    # to_sheet['A1':'N11']
    to_sheet['C3'].value = from_sheet['C3'].value
    to_sheet['C4'].value = from_sheet['C4'].value
    to_sheet['C5'].value = from_sheet['C5'].value
    to_sheet['C6'].value = from_sheet['C6'].value
    to_sheet['C7'].value = from_sheet['C7'].value
    to_sheet['C8'].value = from_sheet['C8'].value
    to_sheet['C9'].value = from_sheet['C9'].value
    to_sheet['C10'].value = from_sheet['C10'].value
    to_sheet['C11'].value = from_sheet['C11'].value
    to_sheet['H3'].value = from_sheet['H3'].value
    to_sheet['H4'].value = from_sheet['H4'].value
    to_sheet['H5'].value = from_sheet['H5'].value
    to_sheet['H6'].value = from_sheet['H6'].value
    to_sheet['H7'].value = from_sheet['H7'].value
    to_sheet['H8'].value = from_sheet['H8'].value
    to_sheet['H9'].value = from_sheet['H9'].value
    to_sheet['H10'].value = from_sheet['H10'].value
    to_sheet['H11'].value = from_sheet['H11'].value


def get_excel_name(well: object, last_depth: float = 0, departure_horiz: float = 0, departure_ver: float = 0) -> str:
    """
    Получаем название файла по параметрам скважины
    """
    # при 0 убираем знак в названии
    departure_horiz = (departure_horiz if departure_horiz != 0.0 else abs(departure_horiz))
    departure_ver = (departure_ver if departure_ver != 0.0 else abs(departure_ver))
    return f"{well.pad_name.field.field_name}_{well.pad_name.pad_name}_{well.well_name}_" \
           f"Скорректированные данные_{last_depth}м (MD) ({departure_horiz}; {departure_ver}).xlsx"
