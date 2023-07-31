""" Новый файл для работы с экселем [удалить]"""
from openpyxl import Workbook


def excel_with_axes():
    """Не используем (можно удалить)"""
    wb = Workbook()
    sheet2 = wb.create_sheet("Данные", 0)
    sheet2.cell(row=1, column=1).value = "Глубина";
    sheet2.cell(row=1, column=2).value = "GX"
    sheet2.cell(row=1, column=3).value = "GY";
    sheet2.cell(row=1, column=4).value = "GZ"
    sheet2.cell(row=1, column=5).value = "BX"
    sheet2.cell(row=1, column=6).value = "BY"
    sheet2.cell(row=1, column=7).value = "BZ"
    for j in range(len(request.POST) - 1):
        if request.POST.get(str(j)) == 'on':
            Data.objects.filter(depth=str(global_data[j][1]),
                                in_statistics=str(global_data[j][0])).update(in_statistics=True)
            for k in range(7):
                if k == 0:
                    sheet2.cell(row=j + 2 - z, column=k + 1).value = (global_data[j])[k + 1]
                else:
                    sheet2.cell(row=j + 2 - z, column=k + 1).value = (global_data[j])[k + 1]
        else:
            Data.objects.filter(depth=str(global_data[j][1]),
                                in_statistics=str(global_data[j][0])).update(in_statistics=False)
            z = z + 1
    wb.remove(wb['Sheet'])
    wb.save("media/data.xlsx")