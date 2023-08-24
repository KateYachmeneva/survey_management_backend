""" Функции по чтению осей с файла и их преобразованию """

import csv
import re
from typing import NoReturn

import lasio
import numpy as np
# import pandas as pd
from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX
from excel_parcer.models import *
from openpyxl.utils import get_column_letter, column_index_from_string

from ..models import List, Device, Data


def toFixed(numObj, digits=0):
    return f"{numObj:.{digits}f}"


def parcing_manually(path, manually_depth, manually_gx, manually_gy, manually_gz, manually_bx, manually_by, manually_bz,
                     manually_import=None) -> list():
    """
    Считываем данные с экселя осей по указанным параметрам. Возращает list с list'ами.
    Лист замеров - лист нижней абстракции это набор параметров глубина - оси.
    """
    try:
        manually_import = int(manually_import)
    except:
        manually_import = None
    data: list = [[], [], [], [], [], [], []]

    if "las" in re.findall(r".(las)", path):
        """Для обработки las"""
        las_file = lasio.read(path)

        for i, name in enumerate(
                [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]):
            if name in las_file.keys():
                data[i] = np.asarray(las_file[name])
        data = np.asarray(data)
        data = data[:, ~np.isnan(data).any(axis=0)]
    # FIXME вынести все условия в функции
    elif "csv" in re.findall(r".(csv)", path):
        """Для обработки csv"""
        with open(path, 'r', newline='') as csvfile:
            i: int = 0
            index_List = [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]
            datareader = csv.reader(csvfile, delimiter=';', quotechar='|')
            for raw in datareader:
                i += 1
                if len(raw) == 0:
                    continue
                for d_index, n_index in enumerate(index_List):
                    try:
                        data[d_index].append(float(raw[int(n_index) - 1]))
                    except Exception as e:
                        pass
    elif "txt" in re.findall(r".(txt)", path):
        """Для обработки txt"""
        index_List = [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]
        start_index = (manually_import if manually_import is not None else 31)
        with open(path, encoding="utf-8") as f:
            for i in range(int(start_index)-1):
                f.readline()  # считываем ненужные строки
            for raw in f.readlines():
                raw_list = raw.replace('\n', '').replace('\t', ',').split(sep=',')
                if len(raw_list) == 0:  # в строке отсутсвует разделитель
                    continue
                for d_index, n_index in enumerate(index_List):
                    try:
                        data[d_index].append(float(raw_list[int(n_index) - 1]))
                    except Exception as e:
                        print(f'Ошибка индексов/неподдерживаемый формат txt (РАЗДЕЛИТЕЛЬ ,) файл с ошибкой: {path}')
                        print(e)
    elif 'sur' in re.findall(r".(sur)", path):
        index_List = [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]
        start_index = (manually_import if manually_import is not None else 2)
        with open(path, encoding="utf-8") as f:
            for i in range(int(start_index)):
                f.readline()
            for raw in f.readlines():
                raw_list = raw.replace('\n', '').split()
                # print(raw_list)
                if len(raw_list) < 7:  # в строке отсутсвует разделитель
                    continue
                for d_index, n_index in enumerate(index_List):
                    data[d_index].append(float(raw_list[int(n_index) - 1]))
    else:
        """Для обработки excel"""
        try:
            wb = load_workbook(filename=path, data_only=True)
        except:
            x2x = XLS2XLSX(path)
            wb = x2x.to_xlsx()

        sheet = wb.active
        for i, name in enumerate(
                [manually_bz, manually_by, manually_bx, manually_gz, manually_gy, manually_gx, manually_depth]):
            colum_id = column_index_from_string(name)
            for row_id in range(int(manually_import), sheet.max_row + 1):
                cell = sheet.cell(row_id, colum_id)
                if cell.value is not None:
                    data[i].append(cell.value)
                else:
                    data[i].append('')

    return tuple(zip(*data[::-1]))


def new_parcing(path):
    try:
        wb = load_workbook(filename=path)
    except:
        x2x = XLS2XLSX(path)
        wb = x2x.to_xlsx()
    sheet = wb.worksheets[0]
    headers = []
    measures = []
    depth = []
    data = []
    GX = []
    GY = []
    GZ = []
    BX = []
    BY = []
    BZ = []
    unique_1 = 0
    unique_2 = 0
    unique_3 = 0
    unique_4 = 0
    unique_5 = 0
    unique_6 = 0
    unique_7 = 0

    for i in sheet.columns:
        for j in range(len(i)):
            for k in List.objects.first().depth.split(';'):
                if i[j].value == k:
                    if unique_1 == 0:
                        unique_1 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    depth.append(z.value)

            for k in List.objects.first().CX.split(';'):
                if i[j].value == k:
                    if unique_2 == 0:
                        unique_2 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    GX.append(z.value)

            for k in (List.objects.first().CY).split(';'):
                if (i[j].value == k):
                    if (unique_3 == 0):
                        unique_3 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    GY.append(z.value)

            for k in (List.objects.first().CZ).split(';'):
                if (i[j].value == k):
                    if (unique_4 == 0):
                        unique_4 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    GZ.append(z.value)

            for k in (List.objects.first().BX).split(';'):
                if (i[j].value == k):
                    if (unique_5 == 0):
                        unique_5 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    BX.append(z.value)

            for k in (List.objects.first().BY).split(';'):
                if (i[j].value == k):
                    if (unique_6 == 0):
                        unique_6 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    BY.append(z.value)

            for k in (List.objects.first().BZ).split(';'):
                if (i[j].value == k):
                    if (unique_7 == 0):
                        unique_7 = 1
                        headers.append(i[j].value)
                        for z in i:
                            if (z.value != None):
                                if (type(z.value) == float):
                                    BZ.append(z.value)
    for i in range(7):
        measures.append("-")

    data.append(headers)
    data.append(measures)
    for i in range(len(depth)):
        result = []
        result.append(depth[i])
        result.append(GX[i])
        result.append(GY[i])
        result.append(GZ[i])
        result.append(BX[i])
        result.append(BY[i])
        result.append(BZ[i])
        # print(result)
        data.append(result)
    return data


def new_measurements(data, name):
    """
    Тоже бы переписать когда-нибудь. Здесь преобразуем данные по указанным правилам.
    Приводим к одному формату.
    """

    result = []

    for i in range(len(data)):
        new_data = []
        for j in range(len(data[i])):
            if '' in data[i]:
                continue
            if j == 0:  # depth[i]
                new_data.append(toFixed((round(float(eval("((data[i])[j])")), 2)), 2))
            if j == 1:  # GX[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).CX)), 7)),
                        7))
            if j == 2:  # GY[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).CY)), 7)),
                        7))
            if j == 3:  # GZ[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).CZ)), 7)),
                        7))
            if j == 4:  # BX[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).BX)), 2)),
                        2))
            if j == 5:  # BY[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).BY)), 2)),
                        2))
            if j == 6:  # BZ[i]
                new_data.append(
                    toFixed((round(
                        float(eval("float(((data[i])[j]))" + Device.objects.get(device_title=name).BZ)), 2)),
                        2))
        result.append(new_data)

    return result


# def dbdata(run):
#     data = [["м", "м/с^2", "м/с^2", "м/с^2", "нТ", "нТ", "нТ"]]
#     for i in Data.objects.filter(run=Run.objects.get(run_title=run)):
#         data.append([i.depth, i.CX, i.CY, i.CZ, i.BX, i.BY, i.BZ])
#     return data


def write_to_bd(data: list[list], run: object) -> NoReturn:
    """
    Записываем преобразованные данные в БД
    """
    update_obj = list()
    for rows in data:
        if len(rows) < 7:  # проверка на наличие всех осей
            continue
        if rows[0] < 0:  # глубина не может быть отрицательной
            continue
        bd_data = Data.objects.get_or_create(depth=rows[0], run=run)
        update_obj.append(bd_data[0])
        bd_data[0].CX = rows[1]
        bd_data[0].CY = rows[2]
        bd_data[0].CZ = rows[3]
        bd_data[0].BX = rows[4]
        bd_data[0].BY = rows[5]
        bd_data[0].BZ = rows[6]
        bd_data[0].in_statistics = True

    Data.objects.bulk_update(update_obj, ["CX", "CY", "CZ", "BX", "BY", "BZ", "in_statistics"])
